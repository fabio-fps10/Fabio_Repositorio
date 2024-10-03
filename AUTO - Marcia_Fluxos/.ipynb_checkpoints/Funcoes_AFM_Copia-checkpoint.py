import pandas as pd
import os
import locale
import openpyxl
from openpyxl.styles import Font, Border, PatternFill, Side
import numpy as np


def Mais_Fluxo(sigla_mes,ultimo_domingo):
    fluxos_atual = pd.read_excel(f"V:\\Vendas\\Acompanhamento Venda Semanais\\{sigla_mes}_2024\\FLUXOS.xlsx")
    
    lista_UE = ['VALE','UNI','SUL','MAIA','CAS','BON','BAR']
    dic_Mais_Fluxo = {}

    for UE in lista_UE:
        fluxo = pd.read_excel(f"C:\\Users\\fabio.souza\\Downloads\\{UE}_{sigla_mes}_{ultimo_domingo}.xls")
        dic_Mais_Fluxo[UE] = fluxo[fluxo[['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2']].notna().all(axis=1)]

    dicionario_dataframes = {}  

    for chave, valor in dic_Mais_Fluxo.items():  
        df = valor 
        df.columns = df.iloc[0]
        df = df[1:]
        df = df[df['Data'] != 'Total']
        dicionario_dataframes[chave] = df.copy()
    

    merge_1 = dicionario_dataframes['VALE'][['Data','Entrada']].merge(dicionario_dataframes['UNI'][['Data','Entrada']],on = 'Data',how ='outer',suffixes = ('_Vale','_UNI')).rename(columns={'Entrada_Vale':'Shopping do Vale','Entrada_UNI':'Unimart Shopping'})
    merge_2 = merge_1.merge(dicionario_dataframes['SUL'][['Data','Entrada']],on = 'Data',how ='outer').rename(columns={'Entrada':'Parque Shopping Sulacap'})
    merge_3 = merge_2.merge(dicionario_dataframes['MAIA'][['Data','Entrada']],on = 'Data',how ='outer').rename(columns={'Entrada':'Parque Shopping Maia'})
    merge_4 = merge_3.merge(dicionario_dataframes['CAS'][['Data','Entrada']],on = 'Data',how ='outer').rename(columns={'Entrada':'Cascavel JL Shopping'})
    merge_5 = merge_4.merge(dicionario_dataframes['BON'][['Data','Entrada']],on = 'Data',how ='outer').rename(columns={'Entrada':'Shopping Bonsucesso'})
    merge_6 = merge_5.merge(dicionario_dataframes['BAR'][['Data','Entrada']],on = 'Data',how ='outer').rename(columns={'Entrada':'Parque Shopping Barueri'})
    
    
    i = 0
    for unidade_economica in list(fluxos_atual['Empreendimento'].unique()):
        try:
            fluxos_atual.loc[(fluxos_atual['Empreendimento'] == unidade_economica) & (fluxos_atual['Mês'] == sigla_mes) & (fluxos_atual['Dia'] < ultimo_domingo + 1), "Mais Fluxo 2024"] = list(merge_6[unidade_economica])
            i = i + 1

        except:
            try:
                fluxos_atual.loc[(fluxos_atual['Empreendimento'] == unidade_economica) & (fluxos_atual['Mês'] == sigla_mes) & (fluxos_atual['Dia'] < qtd_dias_disponiveis + 1), "Mais Fluxo 2024"] = list(merge_6[unidade_economica])
                i = i + 1
                print(i)
            except:
                continue
    if i == 7:
        print("Fluxos Atualizados")
    else:
        print(i)
        print("Algo deu errado no carregamento dos Fluxos")
    
    writer = pd.ExcelWriter(f"V:\\Vendas\\Acompanhamento Venda Semanais\\{sigla_mes}_2024\\FLUXOS.xlsx", engine='xlsxwriter')
    fluxos_atual.to_excel(writer, sheet_name="Planilha1", index=False)
    writer.close()


def tratativa_arquivo(mes_atual,ultimo_domingo):
    meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    dict_cru = {}
   
    # Leitura do arquivo
    df_fluxo_cru = pd.read_excel(f"V:\\Vendas\\Acompanhamento Venda Semanais\\{mes_atual}_2024\\FLUXOS.xlsx")
    
    # defindo colunas
    df_fluxo_V_MF_I8 = df_fluxo_cru[['Empreendimento','Dia','Mês','Mais Fluxo 2019','Mais Fluxo 2022','Mais Fluxo 2023','Mais Fluxo 2024','Iris 8 2019','Iris 8 2022','Iris 8 2023','Iris 8 2024']]
    
    # Acertando os dias para cada mês
    for mes, dados in df_fluxo_V_MF_I8.groupby('Mês'):
        dict_cru[mes] = dados
        if mes == mes_atual:
            dict_cru[mes] = dict_cru[mes][dict_cru[mes]['Dia'] <= ultimo_domingo].groupby(['Empreendimento','Mês']).sum(numeric_only=True)
        else:
            dict_cru[mes] = dict_cru[mes].groupby(['Empreendimento','Mês']).sum(numeric_only=True) 
    
    df_fluxo_V_MF_I8 = pd.concat(dict_cru).reset_index()

    #ordenando os meses
    df_fluxo_V_MF_I8['Mês'] = pd.Categorical(df_fluxo_V_MF_I8['Mês'], categories = meses, ordered=True)
    df_fluxo_V_MF_I8 = df_fluxo_V_MF_I8.sort_values(['Empreendimento','Mês'])

    #Calculando os percentuais
    df_fluxo_V_MF_I8['24/19_(%)'] = ((df_fluxo_V_MF_I8['Mais Fluxo 2024']/df_fluxo_V_MF_I8['Mais Fluxo 2019'])-1)*100
    df_fluxo_V_MF_I8['24/22_(%)'] = ((df_fluxo_V_MF_I8['Mais Fluxo 2024']/df_fluxo_V_MF_I8['Mais Fluxo 2022'])-1)*100
    df_fluxo_V_MF_I8['24/23_(%)'] = ((df_fluxo_V_MF_I8['Mais Fluxo 2024']/df_fluxo_V_MF_I8['Mais Fluxo 2023'])-1)*100

    # Calculando os percentuais para Iris 8
    df_fluxo_V_MF_I8['24/22(%)'] = ((df_fluxo_V_MF_I8['Iris 8 2024']/df_fluxo_V_MF_I8['Iris 8 2022'])-1)*100
    df_fluxo_V_MF_I8['24/23(%)'] = ((df_fluxo_V_MF_I8['Iris 8 2024']/df_fluxo_V_MF_I8['Iris 8 2023'])-1)*100
    
    return df_fluxo_V_MF_I8

def formatar_numeros(mes_atual,ultimo_domingo):
    columns_to_format_percent = ['24/19_(%)', '24/22_(%)', '24/23_(%)', '24/22(%)', '24/23(%)']
    columns_to_format_numbers = ['Mais Fluxo 2019','Mais Fluxo 2022','Mais Fluxo 2023','Mais Fluxo 2024','Iris 8 2022','Iris 8 2023','Iris 8 2024']

    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
    df_fluxo_V_MF_I8 = tratativa_arquivo(mes_atual,ultimo_domingo)
   
    for column in columns_to_format_percent:
        df_fluxo_V_MF_I8[column] = pd.to_numeric(df_fluxo_V_MF_I8[column], errors='coerce')
        df_fluxo_V_MF_I8[column] = df_fluxo_V_MF_I8[column].apply(lambda x: '{:.2f}%'.format(x).replace('.',','))
    
    for column in columns_to_format_numbers:
        df_fluxo_V_MF_I8[column] = pd.to_numeric(df_fluxo_V_MF_I8[column], errors='coerce')
        df_fluxo_V_MF_I8[column] = df_fluxo_V_MF_I8[column].apply(lambda x: locale.format_string('%d', x, grouping=True))

    return df_fluxo_V_MF_I8


def transformar_em_dicionário(mes_atual,ultimo_domingo):
    dict_MFs = {}
    dict_I8 = {}
    df_fluxo_V_MF_I8 = formatar_numeros(mes_atual,ultimo_domingo)

    for empreendimento, dados in df_fluxo_V_MF_I8.groupby('Empreendimento'):
        dict_MFs[empreendimento] = dados[['Empreendimento','Mês','Mais Fluxo 2019','Mais Fluxo 2022','Mais Fluxo 2023','Mais Fluxo 2024','24/19_(%)','24/22_(%)','24/23_(%)']]
        dict_I8[empreendimento] = dados[['Empreendimento','Mês','Iris 8 2022','Iris 8 2023','Iris 8 2024','24/22(%)','24/23(%)']]

    return(dict_MFs,dict_I8)


def Manipulacao_Dados(mes_atual,ultimo_domingo):
    dict_MFs, dict_I8 = transformar_em_dicionário(mes_atual,ultimo_domingo)

    nome_arquivo = f'Fluxos até {ultimo_domingo} de {mes_atual}.xlsx'
    diretorio = r'C:\Users\fabio.souza\Desktop\Day\Fluxo Marcia'
    caminho_arquivo_final = os.path.join(diretorio, nome_arquivo)

    workbook = pd.ExcelWriter(caminho_arquivo_final, engine='xlsxwriter')
    formato_bold = workbook.book.add_format({'bold': True})

    with workbook as writer:
        # ABA 1 - Mais Fluxos
        dict_MFs['Cascavel JL Shopping'].to_excel(writer, sheet_name='Mais Fluxo', startrow=3, startcol=1, index=False)
        dict_MFs['Parque Shopping Barueri'].to_excel(writer, sheet_name='Mais Fluxo', startrow=3, startcol=11, index=False)
        dict_MFs['Parque Shopping Maia'].to_excel(writer, sheet_name='Mais Fluxo', startrow=3, startcol=21, index=False)
        dict_MFs['Shopping Bonsucesso'].to_excel(writer, sheet_name='Mais Fluxo', startrow=19, startcol=1, index=False)
        dict_MFs['Shopping do Vale'].to_excel(writer, sheet_name='Mais Fluxo', startrow=19, startcol=11, index=False)
        dict_MFs['Unimart Shopping'].to_excel(writer, sheet_name='Mais Fluxo', startrow=19, startcol=21, index=False)
        dict_MFs['Parque Shopping Sulacap'].to_excel(writer, sheet_name='Mais Fluxo', startrow=35, startcol=1, index=False)
        
        # ABA 2 - Iris 8
        #dict_I8['Outlet Premium São Paulo'].to_excel(writer, sheet_name='Iris 8', startrow=3, startcol=1, index=False)
        #dict_I8['Outlet Premium Rio de Janeiro'].to_excel(writer, sheet_name='Iris 8', startrow=3, startcol=9, index=False)
        #dict_I8['Outlet Premium Brasilia'].to_excel(writer, sheet_name='Iris 8', startrow=3, startcol=17, index=False)
        #dict_I8['Outlet Premium Salvador'].to_excel(writer, sheet_name='Iris 8', startrow=19, startcol=1, index=False)
        #dict_I8['Outlet Premium Grande São Paulo'].to_excel(writer, sheet_name='Iris 8', startrow=19, startcol=9, index=False)
        #dict_I8['Parque Shopping Barueri'].to_excel(writer, sheet_name='Iris 8', startrow=19, startcol=17, index=False)
        #dict_I8['Parque Shopping Maia'].to_excel(writer, sheet_name='Iris 8', startrow=35, startcol=1, index=False)
        #dict_I8['Parque Shopping Sulacap'].to_excel(writer, sheet_name='Iris 8', startrow=35, startcol=9, index=False)
        #dict_I8['Shopping Bonsucesso'].to_excel(writer, sheet_name='Iris 8', startrow=35, startcol=17, index=False)

        #dict_I8['Outlet Premium Imigrantes'].to_excel(writer, sheet_name='Iris 8', startrow=3, startcol=25, index=False)
        #dict_I8['Cascavel JL Shopping'].to_excel(writer, sheet_name='Iris 8', startrow=19, startcol=25, index=False)
    
        worksheet_mais_fluxos = writer.sheets['Mais Fluxo']
        worksheet_mais_fluxos.write('A2', f'*A comparação para o mês de {mes_atual} inclui dados até o dia {ultimo_domingo}',formato_bold)

        #worksheet_iris_8 = writer.sheets['Iris 8']
        #worksheet_iris_8.write('A2', f'*A comparação para o mês de {mes_atual} inclui dados até o dia {ultimo_domingo}',formato_bold)
  
        for sheet_name in workbook.sheets:
            worksheet = writer.sheets[sheet_name]
            worksheet.hide_gridlines(2)


####################################################################################################################################################################################
            
# Edição Excel
# Importações de módulos

def substituir_e_negrito(sheet, celulas):
    bold_font = Font(bold=True)
    no_border = Border()
    for celula_destino, celula_origem in celulas.items():
        sheet[celula_destino].value = sheet[celula_origem].value
        sheet[celula_destino].font = bold_font
        sheet[celula_destino].border = no_border

def remover_bordas(*sheets):
    no_border = Border()
    for sheet in sheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = no_border

def aplicar_negrito_intervalos(sheet, cell_ranges):
    bold_font = Font(bold=True)
    for cell_range in cell_ranges:
        for row in sheet[cell_range]:
            for cell in row:
                cell.font = bold_font

def aplicar_cor_intervalos(sheet, cell_ranges, cor_hex):
    fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
    for cell_range in cell_ranges:
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = fill

def mudar_cor_e_negrito(sheet, cell_ranges):
    # Definir a cor da fonte para branco e negrito
    font = Font(color="FFFFFF", bold=True)

    # Aplicar a cor da fonte e negrito a cada intervalo de células
    for cell_range in cell_ranges:
        for row in sheet[cell_range]:
            for cell in row:
                cell.font = font
                
            
def dropar_colunas(sheet, colunas):
        for coluna in colunas:
            col_idx = openpyxl.utils.column_index_from_string(coluna)
            sheet.delete_cols(col_idx)


def Edicao_Excel(mes_atual,ultimo_domingo):
    workbook = openpyxl.load_workbook(f"C:\\Users\\fabio.souza\\Desktop\\Day\Fluxo Marcia\\Fluxos até {ultimo_domingo} de {mes_atual}.xlsx")
    sheet_MF = workbook['Mais Fluxo']
    #sheet_I8 = workbook['Iris 8']

    celulas_MF = {'C3': 'B5', 'M3': 'L5', 'W3': 'V5','C19': 'B21', 'M19': 'L21', 'W19': 'V21','C35': 'B37'}

    #celulas_I8 = {'C3': 'B5', 'K3': 'J5', 'S3': 'R5','AA3': 'Z5','C19': 'B21', 'K19': 'J21', 'S19': 'R21','AA19': 'Z21','C35': 'B37', 'K35': 'J37', 'S35': 'R37'}

    cell_ranges_MF = ['C5:C16', 'M5:M16', 'W5:W16', 'C21:C32', 'M21:M32', 'W21:W32', 'C37:C48']
    #cell_ranges_I8 = ['C5:C16', 'K5:K16', 'S5:S16','AA5:AA16', 'C21:C32', 'K21:K32', 'S21:S32','AA21:AA32', 'C37:C48', 'K37:K48', 'S37:S48']

    cell_MF1 = ['B4:J4', 'L4:T4', 'V4:AD4', 'B20:J20', 'L20:T20', 'V20:AD20', 'B36:J36']
    #cell_I81 = ['B4:H4', 'J4:P4', 'R4:X4','Z4:AF4', 'B20:H20', 'J20:P20', 'R20:X20','Z20:AF20', 'B36:H36', 'J36:P36', 'R36:X36']
    cor_hex1 = "538dd5"

    cell_MF2 = ['C3:J3', 'M3:T3', 'W3:AD3', 'C19:J19', 'M19:T19', 'W19:AD19', 'C35:J35']
    #cell_I82 = ['C3:H3', 'K3:P3', 'S3:X3','AA3:AF3', 'C19:H19', 'K19:P19', 'S19:X19','AA19:AF19', 'C35:H35', 'K35:P35', 'S35:X35']
    cor_hex2 = "002060"
  
    substituir_e_negrito(sheet_MF, celulas_MF)
    #substituir_e_negrito(sheet_I8, celulas_I8)
    
    remover_bordas(sheet_MF)

    aplicar_negrito_intervalos(sheet_MF, cell_ranges_MF)
    #aplicar_negrito_intervalos(sheet_I8, cell_ranges_I8)

    aplicar_cor_intervalos(sheet_MF, cell_MF1, cor_hex1)
    #aplicar_cor_intervalos(sheet_I8, cell_I81, cor_hex1)

    aplicar_cor_intervalos(sheet_MF, cell_MF2, cor_hex2)
    #aplicar_cor_intervalos(sheet_I8, cell_I82, cor_hex2)

    mudar_cor_e_negrito(sheet_MF, cell_MF2)
    #mudar_cor_e_negrito(sheet_I8, cell_I82)
    
    dropar_colunas(sheet_MF, ['B', 'K', 'T'])

    workbook.save(f"C:\\Users\\fabio.souza\\Desktop\\Day\\Fluxo Marcia\\Fluxos até {ultimo_domingo} de {mes_atual}.xlsx")

            
def FUNCAO_AUTO_FULXO_MARCIA (mes_atual,ultimo_domingo):
    Mais_Fluxo(mes_atual,ultimo_domingo)
    Manipulacao_Dados(mes_atual,ultimo_domingo)
    Edicao_Excel(mes_atual,ultimo_domingo)
    print('pronto!')
